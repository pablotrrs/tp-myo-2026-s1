#!/usr/bin/env python3
"""
evaluador.py — Sección 4: Evaluación de las estrategias

Lee un archivo .ini de configuración, ejecuta cada modelo contra cada
instancia descubierta en inPath, recolecta métricas desde stdout
(tags [METRIC]) y desde los archivos .out, y genera:
  - Un CSV con todas las métricas (csv_file del .ini)
  - Un XLSX formateado con la tabla que pide la cátedra (negrita en mejores valores)

Uso:
    python evaluador.py config.ini
"""

import sys
import os
import re
import subprocess
import csv
import glob
import time


# ---------------------------------------------------------------------------
# Parseo del .ini
# ---------------------------------------------------------------------------

def parse_ini(path: str) -> dict:
    """Parsea un archivo clave = valor, ignora líneas vacías y comentarios (#)."""
    config = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            key, _, value = line.partition("=")
            config[key.strip()] = value.strip()
    return config


def cargar_config(ini_path: str) -> dict:
    """Valida y estructura la configuración."""
    raw = parse_ini(ini_path)

    config = {
        "csv_file": raw.get("csv_file", "./metrics.csv"),
        "in_path": raw.get("inPath", "./IN/"),
        "threshold": float(raw.get("threshold", 500)),
        "modelos": [],
    }

    for i in range(1, 10):  # soporta hasta 9 modelos por si se agrega algo
        model_key = f"model{i}"
        out_key = f"outPath{i}"
        if model_key not in raw:
            continue
        config["modelos"].append({
            "id": i,
            "script": raw[model_key],
            "out_path": raw.get(out_key, f"./OUT_model{i}/"),
            "nombre": {1: "Salud", 2: "SaludCG", 3: "Challenger"}.get(i, f"Modelo{i}"),
        })

    return config


# ---------------------------------------------------------------------------
# Descubrimiento de instancias
# ---------------------------------------------------------------------------

def descubrir_instancias(in_path: str) -> list:
    """Busca archivos *_pacientes.in y extrae nombres de instancia."""
    pattern = os.path.join(in_path, "*_pacientes.in")
    archivos = glob.glob(pattern)
    instancias = []
    for a in archivos:
        nombre = os.path.basename(a)
        instancia = nombre.replace("_pacientes.in", "")
        instancias.append(instancia)
    return sorted(instancias)


# ---------------------------------------------------------------------------
# Ejecución de modelos
# ---------------------------------------------------------------------------

def ejecutar_modelo(script: str, instancia: str, threshold: float) -> tuple:
    """
    Ejecuta un modelo como subproceso.
    Retorna (stdout, stderr, returncode).
    """
    cmd = [sys.executable, script, instancia, str(threshold)]
    timeout_total = threshold + 120  # margen generoso para setup/teardown

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout_total,
        )
        return result.stdout, result.stderr, result.returncode
    except subprocess.TimeoutExpired:
        return "", "TIMEOUT: el proceso excedió el tiempo límite", -1
    except FileNotFoundError:
        return "", f"ERROR: no se encontró el script {script}", -2


# ---------------------------------------------------------------------------
# Parseo de métricas
# ---------------------------------------------------------------------------

def sanitizar_valor(key: str, value: str) -> str:
    """
    Limpia valores problemáticos:
    - dual_bound de SCIP puede ser 1e+20 (infinito) cuando no hay cota real.
    - Valores None o vacíos se reemplazan por N/A.
    """
    if not value or value.lower() in ("none", ""):
        return "N/A"
    try:
        num = float(value)
        if abs(num) >= 1e+18:  # placeholder de SCIP para "sin cota"
            return "N/A"
    except ValueError:
        pass
    return value


def parsear_metrics_stdout(stdout: str) -> dict:
    """Extrae pares clave=valor de líneas [METRIC] en stdout."""
    metrics = {}
    for match in re.finditer(r"\[METRIC\]\s+(\w+)=(.*)", stdout):
        key = match.group(1)
        value = match.group(2).strip()
        metrics[key] = sanitizar_valor(key, value)
    return metrics


def parsear_objetivo_out(out_file: str) -> str:
    """Lee la línea Z = ... del archivo .out."""
    try:
        with open(out_file) as f:
            for line in f:
                line = line.strip()
                if line.startswith("Z"):
                    # Soporta "Z = 540.0" y "Z=540.0"
                    _, _, val = line.partition("=")
                    return val.strip()
    except FileNotFoundError:
        pass
    return "N/A"


# ---------------------------------------------------------------------------
# Generación de CSV
# ---------------------------------------------------------------------------

# Orden de métricas en la tabla final
METRICAS = [
    ("mejor_objetivo", "Mejor objetivo"),
    ("dual_bound", "Cota dual"),
    ("n_conss", "# restricciones"),
    ("n_vars", "# variables"),
    ("n_vars_last_master", "# variables en últ. maestro"),
]


def mejor_valor(valores: list, metrica_key: str) -> str:
    """
    Dado un conjunto de valores (strings) para una métrica,
    devuelve el mejor según el tipo de métrica.
    - Mejor objetivo: el mayor (queremos maximizar beneficio).
    - Cota dual: la menor (en maximización, una cota superior más baja es más ajustada).
    - # restricciones / variables: el menor (modelo más compacto).
    """
    numericos = []
    for v in valores:
        try:
            numericos.append(float(v))
        except (ValueError, TypeError):
            continue

    if not numericos:
        return None

    if metrica_key == "mejor_objetivo":
        return str(max(numericos))
    else:
        # dual_bound, n_conss, n_vars, n_vars_last_master: menor es mejor
        return str(min(numericos))


def generar_csv(csv_path: str, instancias: list, modelos: list, resultados: dict):
    """Escribe el CSV con el formato Instancia | Métrica | Modelo1 | Modelo2 | ..."""
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        header = ["Instancia", "Métrica"] + [m["nombre"] for m in modelos]
        writer.writerow(header)

        for inst in instancias:
            for metric_key, metric_label in METRICAS:
                row = [inst, metric_label]
                for modelo in modelos:
                    val = resultados.get((inst, modelo["id"]), {}).get(metric_key, "N/A")
                    row.append(val)
                writer.writerow(row)

    print(f"\n[OK] CSV generado: {csv_path}")


# ---------------------------------------------------------------------------
# Generación de XLSX con formato
# ---------------------------------------------------------------------------

def generar_xlsx(xlsx_path: str, instancias: list, modelos: list, resultados: dict):
    """
    Genera un XLSX formateado con la tabla de la sección 4.
    Resalta en negrita el mejor valor por instancia y métrica.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("[WARN] openpyxl no instalado, saltando generación de XLSX.")
        print("       Instalar con: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluación"

    # Estilos
    font_header = Font(name="Arial", bold=True, size=11)
    font_normal = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", bold=True, size=10)
    font_inst = Font(name="Arial", bold=True, size=10)
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font_header_white = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header
    headers = ["Instancia", "Métrica"] + [m["nombre"] for m in modelos]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font_header_white
        cell.fill = fill_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Datos
    row_num = 2
    for inst in instancias:
        for idx, (metric_key, metric_label) in enumerate(METRICAS):
            # Columna Instancia (solo en la primera fila del grupo)
            cell_inst = ws.cell(row=row_num, column=1)
            if idx == 0:
                cell_inst.value = inst
                cell_inst.font = font_inst
            cell_inst.border = thin_border

            # Columna Métrica
            cell_met = ws.cell(row=row_num, column=2, value=metric_label)
            cell_met.font = font_normal
            cell_met.border = thin_border

            # Valores de cada modelo
            valores = []
            for modelo in modelos:
                val = resultados.get((inst, modelo["id"]), {}).get(metric_key, "N/A")
                valores.append(val)

            mejor = mejor_valor(valores, metric_key)

            for col_offset, val in enumerate(valores):
                cell = ws.cell(row=row_num, column=3 + col_offset)
                # Intentar escribir como número
                try:
                    cell.value = float(val)
                    cell.number_format = "#,##0.00" if "." in val else "#,##0"
                except (ValueError, TypeError):
                    cell.value = val

                # Negrita si es el mejor
                if mejor is not None and val != "N/A":
                    try:
                        if abs(float(val) - float(mejor)) < 1e-6:
                            cell.font = font_bold
                        else:
                            cell.font = font_normal
                    except (ValueError, TypeError):
                        cell.font = font_normal
                else:
                    cell.font = font_normal

                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            row_num += 1

    # Merge celdas de instancia para legibilidad
    row_num = 2
    for inst in instancias:
        if len(METRICAS) > 1:
            ws.merge_cells(
                start_row=row_num, start_column=1,
                end_row=row_num + len(METRICAS) - 1, end_column=1,
            )
            ws.cell(row=row_num, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        row_num += len(METRICAS)

    # Ajustar anchos
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    for i, _ in enumerate(modelos):
        col_letter = chr(ord("C") + i)
        ws.column_dimensions[col_letter].width = 18

    wb.save(xlsx_path)
    print(f"[OK] XLSX generado: {xlsx_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Uso: python evaluador.py <archivo.ini>")
        print("Ejemplo: python evaluador.py config.ini")
        sys.exit(1)

    ini_path = sys.argv[1]
    if not os.path.exists(ini_path):
        print(f"[ERROR] No se encontró {ini_path}")
        sys.exit(1)

    config = cargar_config(ini_path)
    instancias = descubrir_instancias(config["in_path"])

    if not instancias:
        print(f"[ERROR] No se encontraron instancias en {config['in_path']}")
        sys.exit(1)

    if not config["modelos"]:
        print("[ERROR] No se encontraron modelos en el .ini")
        sys.exit(1)

    print("=" * 70)
    print("EVALUADOR — Sección 4")
    print("=" * 70)
    print(f"  Instancias: {instancias}")
    print(f"  Modelos:    {[m['nombre'] for m in config['modelos']]}")
    print(f"  Threshold:  {config['threshold']}s")
    print(f"  CSV:        {config['csv_file']}")
    print("=" * 70)

    resultados = {}  # (instancia, modelo_id) -> dict de métricas

    for inst in instancias:
        for modelo in config["modelos"]:
            print(f"\n{'─'*70}")
            print(f"  {modelo['nombre']}  ×  {inst}")
            print(f"{'─'*70}")

            t0 = time.time()
            stdout, stderr, rc = ejecutar_modelo(
                modelo["script"], inst, config["threshold"]
            )
            elapsed = time.time() - t0

            if rc < 0:
                print(f"  [FAIL] {stderr}")
            else:
                print(f"  [OK] Terminó en {elapsed:.1f}s (rc={rc})")

            # Parsear métricas del stdout
            metrics = parsear_metrics_stdout(stdout)

            # Parsear objetivo del .out
            out_file = os.path.join(modelo["out_path"], f"{inst}.out")
            objetivo = parsear_objetivo_out(out_file)
            metrics["mejor_objetivo"] = objetivo

            # Si no hay dual_bound en [METRIC], marcar N/A
            metrics.setdefault("dual_bound", "N/A")
            metrics.setdefault("n_conss", "N/A")
            metrics.setdefault("n_vars", "N/A")
            metrics.setdefault("n_vars_last_master", "-")

            resultados[(inst, modelo["id"])] = metrics

            print(f"  → Objetivo: {objetivo}")
            for k, v in metrics.items():
                if k != "mejor_objetivo":
                    print(f"  → {k}: {v}")

    # Generar salida
    generar_csv(config["csv_file"], instancias, config["modelos"], resultados)

    xlsx_path = config["csv_file"].replace(".csv", ".xlsx")
    generar_xlsx(xlsx_path, instancias, config["modelos"], resultados)

    print("\n" + "=" * 70)
    print("EVALUACIÓN COMPLETA")
    print("=" * 70)


if __name__ == "__main__":
    main()